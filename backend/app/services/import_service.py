from __future__ import annotations

import csv
import io
import json
from collections.abc import Iterable
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.exceptions import AppError
from app.db.models import Company, Employee, TimeEntrySource, User
from app.repositories.employee_repository import EmployeeRepository
from app.schemas.employee import EmployeeCreateRequest
from app.schemas.time_entry import (
    ImportRowError,
    ImportRowSuccess,
    TimeEntryCreateRequest,
    TimeEntryImportResponse,
)
from app.services.audit_service import record_audit_event
from app.services.time_entry_service import create_time_entry_with_calculation


employee_repository = EmployeeRepository()

HEADER_SYNONYMS: dict[str, tuple[str, ...]] = {
    "employee_email": ("employee_email", "correo", "email", "correo_empleado"),
    "employee_code": ("employee_code", "codigo", "codigo_empleado"),
    "employee_name": ("employee_name", "nombre", "empleado", "nombre_empleado"),
    "area": ("area", "departamento"),
    "position": ("position", "cargo"),
    "age": ("age", "edad"),
    "base_salary": ("base_salary", "salario_base", "salario"),
    "work_date": ("work_date", "fecha", "fecha_turno"),
    "check_in": ("check_in", "hora_entrada", "entrada"),
    "check_out": ("check_out", "hora_salida", "salida"),
    "is_holiday": ("is_holiday", "festivo", "es_festivo"),
    "is_sunday": ("is_sunday", "dominical", "es_dominical"),
    "weekly_accumulated_hours": (
        "weekly_accumulated_hours",
        "acumulado_semanal_horas",
        "acumulado_semana",
    ),
    "weekly_hours": ("weekly_hours", "horas_semanales", "jornada_semanal"),
    "work_days_per_week": ("work_days_per_week", "dias_laborales_semana", "dias_semana"),
}


def normalize_header(value: str) -> str:
    return (
        value.strip()
        .lower()
        .replace(" ", "_")
        .replace("-", "_")
        .replace(".", "")
        .replace("/", "_")
    )


def resolve_mapping(
    headers: list[str], explicit_mapping: dict[str, str] | None = None
) -> dict[str, str]:
    normalized_headers = {normalize_header(header): header for header in headers}
    mapping: dict[str, str] = {}

    if explicit_mapping:
        for field_name, header in explicit_mapping.items():
            mapping[field_name] = header

    for field_name, candidates in HEADER_SYNONYMS.items():
        if field_name in mapping:
            continue
        for candidate in candidates:
            normalized = normalize_header(candidate)
            if normalized in normalized_headers:
                mapping[field_name] = normalized_headers[normalized]
                break

    return mapping


def parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    normalized = str(value or "").strip().lower()
    return normalized in {"1", "true", "si", "yes", "y", "x"}


def parse_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    normalized = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(normalized, fmt).date()
        except ValueError:
            continue

    raise AppError(
        f"Fecha no valida: {normalized}. Usa formatos como YYYY-MM-DD o DD/MM/YYYY.",
        code="invalid_date",
    )


def parse_time(value: Any) -> Any:
    if hasattr(value, "strftime"):
        return value

    normalized = str(value).strip()
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(normalized, fmt).time()
        except ValueError:
            continue

    raise AppError(
        f"Hora no valida: {normalized}. Usa formatos como HH:MM.",
        code="invalid_time",
    )


def read_rows(filename: str, payload: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    extension = filename.lower().split(".")[-1]
    if extension == "csv":
        text = payload.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return list(reader.fieldnames or []), list(reader)

    if extension in {"xlsx", "xlsm"}:
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        headers = [str(cell or "").strip() for cell in rows[0]]
        data_rows = [
            {headers[index]: row[index] for index in range(len(headers))}
            for row in rows[1:]
            if any(value not in (None, "") for value in row)
        ]
        return headers, data_rows

    raise AppError("Formato de archivo no soportado. Usa CSV o XLSX.", code="invalid_file")


def upsert_employee_from_row(
    db: Session,
    *,
    actor: User,
    company: Company,
    row: dict[str, Any],
    mapping: dict[str, str],
    create_missing: bool,
) -> Employee:
    email = row.get(mapping.get("employee_email", ""), None)
    code = row.get(mapping.get("employee_code", ""), None)
    name = row.get(mapping.get("employee_name", ""), None)

    employee = employee_repository.get_by_identity(
        db,
        company_id=company.id,
        email=str(email).strip().lower() if email else None,
        external_code=str(code).strip() if code else None,
        full_name=str(name).strip() if name else None,
    )
    if employee:
        return employee

    if not create_missing:
        raise AppError(
            "No se encontro un empleado existente para la fila y create_missing_employees esta desactivado.",
            code="employee_not_found",
        )

    payload = EmployeeCreateRequest(
        external_code=str(code).strip() if code else None,
        full_name=str(name or "").strip(),
        email=str(email).strip().lower() if email else None,
        position=str(row.get(mapping.get("position", ""), "") or "Sin cargo"),
        area=str(row.get(mapping.get("area", ""), "") or "Sin area"),
        age=int(row.get(mapping.get("age", ""), 18) or 18),
        base_salary=float(row.get(mapping.get("base_salary", ""), 0) or 0),
        weekly_hours=float(row.get(mapping.get("weekly_hours", ""), 42) or 42),
        work_days_per_week=int(row.get(mapping.get("work_days_per_week", ""), 5) or 5),
        status="activo",
    )
    if payload.base_salary <= 0 or not payload.full_name:
        raise AppError(
            "No hay suficientes datos para crear automaticamente el empleado de la fila.",
            code="employee_creation_failed",
        )

    return employee_repository.create(
        db,
        company_id=company.id,
        external_code=payload.external_code,
        full_name=payload.full_name,
        email=payload.email,
        position=payload.position,
        area=payload.area,
        age=payload.age,
        base_salary=payload.base_salary,
        weekly_hours=payload.weekly_hours,
        work_days_per_week=payload.work_days_per_week,
        status=payload.status,
        created_by_id=actor.id,
        updated_by_id=actor.id,
    )


def import_time_entries(
    db: Session,
    *,
    actor: User,
    company: Company,
    filename: str,
    payload: bytes,
    explicit_mapping: dict[str, str] | None = None,
    create_missing_employees: bool = False,
) -> TimeEntryImportResponse:
    headers, rows = read_rows(filename, payload)
    mapping = resolve_mapping(headers, explicit_mapping)
    required_fields = ("work_date", "check_in", "check_out")
    missing_fields = [field_name for field_name in required_fields if field_name not in mapping]
    if missing_fields:
        raise AppError(
            f"Faltan columnas obligatorias para importar: {', '.join(missing_fields)}.",
            code="missing_import_columns",
        )

    successes: list[ImportRowSuccess] = []
    errors: list[ImportRowError] = []

    for row_number, row in enumerate(rows, start=2):
        try:
            with db.begin_nested():
                employee = upsert_employee_from_row(
                    db,
                    actor=actor,
                    company=company,
                    row=row,
                    mapping=mapping,
                    create_missing=create_missing_employees,
                )
                time_entry_payload = TimeEntryCreateRequest(
                    employee_id=employee.id,
                    work_date=parse_date(row[mapping["work_date"]]),
                    check_in=parse_time(row[mapping["check_in"]]),
                    check_out=parse_time(row[mapping["check_out"]]),
                    is_holiday=parse_bool(row.get(mapping.get("is_holiday", ""), False)),
                    is_sunday=parse_bool(row.get(mapping.get("is_sunday", ""), False)),
                    weekly_accumulated_hours=float(
                        row.get(mapping.get("weekly_accumulated_hours", ""), 0) or 0
                    ),
                    notes=f"Importado desde {filename}",
                )
                entry = create_time_entry_with_calculation(
                    db,
                    actor=actor,
                    employee=employee,
                    payload=time_entry_payload,
                    company_settings=company.settings_json or {},
                    source=TimeEntrySource.IMPORT,
                )
            successes.append(
                ImportRowSuccess(
                    row_number=row_number,
                    time_entry_id=entry.id,
                    employee_id=employee.id,
                    employee_name=employee.full_name,
                )
            )
        except Exception as exc:  # noqa: BLE001
            errors.append(
                ImportRowError(
                    row_number=row_number,
                    row_data={key: str(value) if value is not None else "" for key, value in row.items()},
                    reason=str(exc),
                )
            )

    audit_event = record_audit_event(
        db,
        actor=actor,
        action="time_entry.imported",
        entity_type="time_entry_import",
        metadata={
            "filename": filename,
            "total_rows": len(rows),
            "successful_rows": len(successes),
            "rejected_rows": len(errors),
            "mapping": mapping,
            "errors": [error.model_dump() for error in errors],
        },
    )

    return TimeEntryImportResponse(
        audit_event_id=audit_event.id,
        total_rows=len(rows),
        successful_rows=len(successes),
        rejected_rows=len(errors),
        successes=successes,
        errors=errors,
        error_report_download_url=(
            f"/api/v1/time-entries/imports/{audit_event.id}/errors.csv" if errors else None
        ),
    )


def export_import_errors_csv(errors: Iterable[dict[str, Any]]) -> str:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["row_number", "reason", "row_data"])
    for error in errors:
        writer.writerow(
            [
                error.get("row_number"),
                error.get("reason"),
                json.dumps(error.get("row_data", {}), ensure_ascii=True),
            ]
        )
    return buffer.getvalue()
